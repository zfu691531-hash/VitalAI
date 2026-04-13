import io
import uuid
from datetime import datetime
from typing import List, Optional, Union
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment


def dict_list_to_excel(
    data: List[dict],
    output_path: Optional[str] = None,
    sheet_name: str = "Sheet1",
    field_name_map: Optional[dict] = None,
    exclude_fields: Optional[List[str]] = None,
    include_all_fields: bool = False
) -> Union[bytes, None]:
    """
    将 List[dict] 数据转换为 Excel 文件

    Args:
        data: 字典列表，每个字典代表一行数据
        output_path: 输出文件路径，如果为 None 则返回字节流（用于 FastAPI 接口）
        sheet_name: 工作表名称，默认为 "Sheet1"
        field_name_map: 字段名映射字典，key为原字段名，value为显示的中文列名（可选）
        exclude_fields: 需要排除的字段列表（可选）
        include_all_fields: 是否包含 field_name_map 中定义的所有字段，即使某些字段在数据中不存在（默认 False）

    Returns:
        如果 output_path 为 None，返回 Excel 文件的字节流（bytes）
        如果 output_path 有值，保存到文件并返回 None

    Example:
        # 示例1：直接返回字节流（用于 FastAPI）
        data = [
            {"name": "张三", "age": 25, "city": "北京"},
            {"name": "李四", "age": 30, "city": "上海"}
        ]
        excel_bytes = dict_list_to_excel(data, field_name_map={"name": "姓名", "age": "年龄", "city": "城市"})
        with open('output.xlsx', 'wb') as f:
            f.write(excel_bytes)

        # 示例2：保存到文件
        dict_list_to_excel(data, output_path='output.xlsx')

        # 示例3：在 FastAPI 中使用，确保所有列都显示
        @app.get("/export")
        async def export_data():
            data = query_data()  # 查询数据
            excel_bytes = dict_list_to_excel(data, field_name_map=all_fields_map, include_all_fields=True)
            return Response(
                content=excel_bytes,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": "attachment; filename=data.xlsx"}
            )
    """
    if not data:
        raise ValueError("数据不能为空")

    # 处理字段映射和排除字段
    if exclude_fields is None:
        exclude_fields = []

    # 获取所有字段名
    if include_all_fields and field_name_map:
        # 使用 field_name_map 中定义的所有字段
        headers = [k for k in field_name_map.keys() if k not in exclude_fields]
    else:
        # 取所有字典的并集
        all_headers = set()
        for row in data:
            all_headers.update(row.keys())
        headers = [h for h in all_headers if h not in exclude_fields]

    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # 设置表头样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    # 写入表头（使用字段名映射）
    for col_idx, header in enumerate(headers, 1):
        display_header = field_name_map.get(header, header) if field_name_map else header
        cell = ws.cell(row=1, column=col_idx, value=display_header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # 写入数据
    for row_idx, row_data in enumerate(data, start=2):
        for col_idx, header in enumerate(headers, 1):
            value = row_data.get(header, "")
            ws.cell(row=row_idx, column=col_idx, value=value)

    # 自动调整列宽
    for col_idx, header in enumerate(headers, 1):
        display_header = field_name_map.get(header, header) if field_name_map else header
        max_length = len(str(display_header))
        for row_data in data:
            value = str(row_data.get(header, ""))
            # 考虑中文字符，长度乘以1.5
            length = len(value)
            # 计算实际显示长度（中文字符占更多宽度）
            actual_length = sum(1.5 if '\u4e00' <= char <= '\u9fff' else 1 for char in value)
            max_length = max(max_length, actual_length)
        ws.column_dimensions[chr(64 + col_idx)].width = min(max_length + 2, 50)

    # 保存到 BytesIO 或文件
    if output_path:
        wb.save(output_path)
        return None
    else:
        # 保存到内存中的 BytesIO 对象
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        return excel_buffer.getvalue()


def dict_list_to_excel_stream(data: List[dict], sheet_name: str = "Sheet1") -> io.BytesIO:
    """
    将 List[dict] 数据转换为 Excel 文件流（返回 BytesIO 对象）

    Args:
        data: 字典列表，每个字典代表一行数据
        sheet_name: 工作表名称，默认为 "Sheet1"

    Returns:
        BytesIO 对象，可直接用于 FastAPI 的 FileResponse 或 StreamingResponse

    Example:
        @app.get("/export")
        async def export_data():
            data = query_data()  # 查询数据
            excel_stream = dict_list_to_excel_stream(data)
            return StreamingResponse(
                io.BytesIO(excel_stream.getvalue()),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": "attachment; filename=data.xlsx"}
            )
    """
    excel_bytes = dict_list_to_excel(data, sheet_name=sheet_name)
    return io.BytesIO(excel_bytes)


def excel_to_dict_list(
    file_path: Optional[str] = None,
    file_obj: Optional[io.BytesIO] = None,
    sheet_name: Optional[str] = None,
    field_name_map: Optional[dict] = None,
    skip_rows: int = 1,
    skip_empty_rows: bool = True
) -> List[dict]:
    """
    读取 Excel 文件并转换为字典列表

    Args:
        file_path: Excel 文件路径（可选，与 file_obj 二选一）
        file_obj: Excel 文件对象（BytesIO 或 File-like），优先使用 file_obj（可选）
        sheet_name: 工作表名称，默认读取第一个工作表
        field_name_map: 字段名映射字典，key为Excel列名（中文），value为目标字段名（英文）
        skip_rows: 跳过的行数（表头），默认跳过1行
        skip_empty_rows: 是否跳过空行，默认为 True

    Returns:
        字典列表，每个字典代表一行数据

    Example:
        # 示例1：从文件路径读取
        data = excel_to_dict_list(file_path='data.xlsx')

        # 示例2：从文件对象读取
        from fastapi import UploadFile
        file_obj = await upload_file.read()
        data = excel_to_dict_list(file_obj=io.BytesIO(file_obj))

        # 示例3：使用字段名映射（中文 -> 英文）
        data = excel_to_dict_list(
            file_path='data.xlsx',
            field_name_map={'姓名': 'name', '年龄': 'age', '城市': '城市'}
        )
        # 返回: [{'name': '张三', 'age': 25, 'city': '北京'}, ...]
    """
    if not file_path and not file_obj:
        raise ValueError("必须提供 file_path 或 file_obj 参数")

    try:
        # 优先使用 file_obj，否则使用 file_path
        if file_obj:
            file_obj.seek(0)  # 重置指针
            wb = load_workbook(file_obj)
        else:
            wb = load_workbook(file_path)

        # 获取工作表
        if sheet_name:
            ws = wb[sheet_name]
        else:
            ws = wb.active

        data_list = []

        # 获取表头
        headers = []
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col).value
            if header:
                # 如果提供了字段名映射，进行转换
                if field_name_map and header in field_name_map:
                    header = field_name_map[header]
                headers.append(header)

        # 从第二行开始读取数据
        for row_idx in range(skip_rows + 1, ws.max_row + 1):
            row_data = {}

            # 读取每列数据
            for col_idx, header in enumerate(headers, start=1):
                cell_value = ws.cell(row=row_idx, column=col_idx).value

                # 处理不同类型的单元格值
                if isinstance(cell_value, datetime):
                    cell_value = cell_value.strftime('%Y-%m-%d %H:%M:%S')
                elif cell_value is None:
                    cell_value = ""

                row_data[header] = cell_value

            # 跳过空行
            if skip_empty_rows and not any(row_data.values()):
                continue

            data_list.append(row_data)

        return data_list

    except Exception as e:
        raise ValueError(f"读取 Excel 文件失败：{str(e)}")
