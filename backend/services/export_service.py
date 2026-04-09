# -*- coding: utf-8 -*-
"""
导出服务模块
============
- generate_student_excel: 生成学生信息Excel文件
- generate_score_excel: 生成成绩Excel文件
- download_template: 生成标准导入模板（学生/成绩）
- parse_import_excel: 解析上传的Excel文件，返回数据列表与校验错误

注意：本模块为纯工具模块，不访问数据库，字段转换由调用者（M6/M9）负责。
"""

import uuid
import openpyxl
from io import BytesIO
from typing import Tuple, List, Dict
from fastapi import UploadFile
from fastapi.responses import StreamingResponse
from utils.logger import logger

# 允许的文件扩展名
ALLOWED_EXTENSIONS = {".xlsx", ".xls"}
# 最大文件大小 5MB
MAX_FILE_SIZE = 5 * 1024 * 1024
# 单次导出最大行数
MAX_EXPORT_ROWS = 5000

# 学生模板表头（带星号为必填）
STUDENT_HEADERS = ["学号*", "姓名*", "性别*", "年龄", "班级名称*", "联系方式", "特长", "标签"]
# 成绩模板表头（带星号为必填）
SCORE_HEADERS = ["学号*", "班级名称*", "考试批次*", "科目*", "分数*"]

# 必填表头映射（去除星号）
REQUIRED_STUDENT_HEADERS = ["学号", "姓名", "性别", "班级名称"]
REQUIRED_SCORE_HEADERS = ["学号", "班级名称", "考试批次", "科目", "分数"]


def generate_student_excel(rows: List[Dict]) -> StreamingResponse:
    """
    生成学生信息Excel文件
    
    Args:
        rows: 学生数据列表，每个字典包含学生信息
              [{"学号": "001", "姓名": "张三", ...}, ...]
    
    Returns:
        StreamingResponse: Excel文件流式响应
    """
    if len(rows) > MAX_EXPORT_ROWS:
        raise ValueError(f"导出行数超过限制（最大{MAX_EXPORT_ROWS}行），请分批导出")
    
    # 清理表头（去除星号）
    headers = [h.rstrip("*") for h in STUDENT_HEADERS]
    
    try:
        # 判断是否使用write_only模式
        use_write_only = len(rows) > 1000
        
        if use_write_only:
            workbook = openpyxl.Workbook(write_only=True)
            ws = workbook.create_sheet("学生信息")
        else:
            workbook = openpyxl.Workbook()
            ws = workbook.active
            ws.title = "学生信息"
        
        # 写入表头
        if use_write_only:
            ws.append(headers)
        else:
            ws.append(headers)
            # 设置表头加粗
            for cell in ws[1]:
                cell.font = openpyxl.styles.Font(bold=True)
            # 设置列宽
            column_widths = [15, 12, 8, 8, 20, 15, 20, 30]
            for i, width in enumerate(column_widths, 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
        
        # 写入数据行
        for row in rows:
            row_data = [row.get(h, "") for h in headers]
            ws.append(row_data)
        
        # 保存到内存
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        
        logger.info(f"生成学生Excel成功，共{len(rows)}行")
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=students.xlsx"}
        )
        
    except Exception as e:
        logger.error(f"生成学生Excel失败: {e}")
        raise


def generate_score_excel(rows: List[Dict]) -> StreamingResponse:
    """
    生成成绩Excel文件
    
    Args:
        rows: 成绩数据列表，每个字典包含成绩信息
              [{"学号": "001", "班级名称": "一年级一班", ...}, ...]
    
    Returns:
        StreamingResponse: Excel文件流式响应
    """
    if len(rows) > MAX_EXPORT_ROWS:
        raise ValueError(f"导出行数超过限制（最大{MAX_EXPORT_ROWS}行），请分批导出")
    
    # 清理表头（去除星号）
    headers = [h.rstrip("*") for h in SCORE_HEADERS]
    
    try:
        use_write_only = len(rows) > 1000
        
        if use_write_only:
            workbook = openpyxl.Workbook(write_only=True)
            ws = workbook.create_sheet("成绩信息")
        else:
            workbook = openpyxl.Workbook()
            ws = workbook.active
            ws.title = "成绩信息"
        
        # 写入表头
        if use_write_only:
            ws.append(headers)
        else:
            ws.append(headers)
            for cell in ws[1]:
                cell.font = openpyxl.styles.Font(bold=True)
            column_widths = [15, 20, 15, 12, 10]
            for i, width in enumerate(column_widths, 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
        
        # 写入数据行
        for row in rows:
            row_data = [row.get(h, "") for h in headers]
            ws.append(row_data)
        
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        
        logger.info(f"生成成绩Excel成功，共{len(rows)}行")
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=scores.xlsx"}
        )
        
    except Exception as e:
        logger.error(f"生成成绩Excel失败: {e}")
        raise


def download_template(template_type: str) -> StreamingResponse:
    """
    生成标准导入模板（学生/成绩）
    
    Args:
        template_type: 模板类型，"student" 或 "score"
    
    Returns:
        StreamingResponse: Excel模板文件流式响应
    """
    try:
        workbook = openpyxl.Workbook()
        ws = workbook.active
        
        if template_type == "student":
            ws.title = "学生导入模板"
            headers = STUDENT_HEADERS
            # 示例数据
            example_data = ["S001", "张三", "男", "18", "一年级一班", "13800138000", "篮球", "活泼,积极"]
            filename = "student_import_template.xlsx"
        elif template_type == "score":
            ws.title = "成绩导入模板"
            headers = SCORE_HEADERS
            example_data = ["S001", "一年级一班", "期中考试", "数学", "95.5"]
            filename = "score_import_template.xlsx"
        else:
            raise ValueError(f"不支持的模板类型: {template_type}")
        
        # 写入表头
        ws.append(headers)
        for cell in ws[1]:
            cell.font = openpyxl.styles.Font(bold=True)
        
        # 写入示例数据
        ws.append(example_data)
        
        # 设置列宽
        for i in range(1, len(headers) + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 15
        
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        
        logger.info(f"生成{template_type}模板成功")
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except Exception as e:
        logger.error(f"生成模板失败: {e}")
        raise


def parse_import_excel(
    file: UploadFile,
    template_type: str
) -> Tuple[List[Dict], List[Dict]]:
    """
    解析上传的Excel文件，返回数据列表与校验错误
    
    注意：本函数仅解析Excel返回原始字符串数据，
    字段转换（如班级名称→class_id）由调用者负责。
    
    Args:
        file: 上传的文件对象
        template_type: 模板类型，"student" 或 "score"
    
    Returns:
        tuple: (有效数据列表, 错误列表)
               valid_data: [{"学号": "001", "姓名": "张三", ...}, ...]
               error_list: [{"row": 3, "message": "学号不能为空"}, ...]
    """
    error_list: List[Dict] = []
    
    # 安全检查：文件扩展名
    file_ext = "." + file.filename.rsplit(".", 1)[-1].lower() if "." in file.filename else ""
    if file_ext not in ALLOWED_EXTENSIONS:
        error_list.append({"row": 0, "message": f"不支持的文件类型: {file_ext}，仅支持 .xlsx 和 .xls"})
        return [], error_list
    
    try:
        # 安全检查：文件大小
        content = file.file.read()
        if len(content) > MAX_FILE_SIZE:
            error_list.append({"row": 0, "message": f"文件大小超过限制（最大5MB）"})
            return [], error_list
        
        # 使用uuid命名临时文件
        temp_filename = f"{uuid.uuid4()}{file_ext}"
        logger.debug(f"解析Excel文件: {temp_filename}")
        
        # 解析Excel
        workbook = openpyxl.load_workbook(BytesIO(content))
        ws = workbook.active
        
        # 获取表头（清理星号）
        header_row = [str(cell.value).rstrip("*") if cell.value else "" for cell in ws[1]]
        
        # 确定必填表头
        if template_type == "student":
            required_headers = REQUIRED_STUDENT_HEADERS
        elif template_type == "score":
            required_headers = REQUIRED_SCORE_HEADERS
        else:
            error_list.append({"row": 0, "message": f"不支持的模板类型: {template_type}"})
            return [], error_list
        
        # 校验必填表头是否存在
        missing_headers = [h for h in required_headers if h not in header_row]
        if missing_headers:
            error_list.append({"row": 0, "message": f"缺少必填列: {', '.join(missing_headers)}"})
            return [], error_list
        
        # 解析数据行
        valid_data: List[Dict] = []
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            # 跳过空行
            if all(cell is None or str(cell).strip() == "" for cell in row):
                continue
            
            row_data = {}
            row_errors = []
            
            for col_idx, header in enumerate(header_row):
                if col_idx < len(row):
                    value = row[col_idx]
                    row_data[header] = str(value).strip() if value is not None else ""
                else:
                    row_data[header] = ""
            
            # 校验必填字段
            for required in required_headers:
                if not row_data.get(required):
                    row_errors.append(f"{required}不能为空")
            
            if row_errors:
                error_list.append({"row": row_num, "message": "；".join(row_errors)})
            else:
                valid_data.append(row_data)
        
        logger.info(f"解析Excel完成，有效数据{len(valid_data)}行，错误{len(error_list)}行")
        
        return valid_data, error_list
        
    except openpyxl.utils.exceptions.InvalidFileException:
        error_list.append({"row": 0, "message": "文件格式错误，无法解析"})
        return [], error_list
    except Exception as e:
        logger.error(f"解析Excel失败: {e}")
        error_list.append({"row": 0, "message": "文件解析异常，请检查文件格式"})
        return [], error_list
